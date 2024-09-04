import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
import pandas as pd
from datetime import datetime
import tkinter.messagebox as messagebox
import requests
import pyodbc
from openpyxl import load_workbook 
import os


def create_app():
    api_url= 'http://localhost:5005/api/Meiji/get_data'
    factory_name= '明治'
    value_bolean = 1
    root = tk.Tk()
    button_bg_color = "#d3d3d3" 

    ##d3d3d3 // mau xam
    #4da6ff //xanh duong
    #FF5722 mau cam

    root.title("生産管理版")


    # Tạo nhãn cho tiêu đề
    title_label = tk.Label(root, text="生　産　管　理　版", font=("Helvetica", 20, "bold"))
    title_label.pack(pady=10)

    root.tk.call("source", "azure.tcl")
    root.tk.call("set_theme", "dark")

    # Tạo nhãn cho thông tin chung
    info_frame = tk.Frame(root)
    info_frame.pack(pady=10) 

    tk.Label(info_frame, text="車種").grid(row=0, column=0, padx=10)
    car_type_combobox = ttk.Combobox(info_frame)
    car_type_combobox.grid(row=0, column=1, padx=10)    



    tk.Label(info_frame, text="工場").grid(row=0, column=2, padx=10)
    # Tạo combobox cho 車種
    factory_types = ["平子", "明治川"]
    factory_type_combobox = ttk.Combobox(info_frame, values=factory_types)
    factory_type_combobox.grid(row=0, column=3, padx=10)
    factory_type_combobox.current(1) 

    # Tạo DateEntry cho 日付
    single_date_label = tk.Label(info_frame, text="日付")
    single_date_entry = DateEntry(info_frame, width=12, background='darkblue', foreground='white', borderwidth=2)

    range_date_label = tk.Label(info_frame, text="開始日付")
    start_date_entry = DateEntry(info_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
    end_date_label = tk.Label(info_frame, text="終了日付")
    end_date_entry = DateEntry(info_frame, width=12, background='darkblue', foreground='white', borderwidth=2)

    # Đặt vị trí mặc định cho ngày đơn
    single_date_label.grid(row=0, column=4, padx=10)
    single_date_entry.grid(row=0, column=5, padx=10)

    def toggle_date_mode():
        if single_date_label.winfo_ismapped():
            single_date_label.grid_remove()
            single_date_entry.grid_remove()
            range_date_label.grid(row=0, column=4, padx=10)
            start_date_entry.grid(row=0, column=5, padx=10)
            end_date_label.grid(row=0, column=6, padx=10)
            end_date_entry.grid(row=0, column=7, padx=10)
        else:
            range_date_label.grid_remove()
            start_date_entry.grid_remove()
            end_date_label.grid_remove()
            end_date_entry.grid_remove()
            single_date_label.grid(row=0, column=4, padx=10)
            single_date_entry.grid(row=0, column=5, padx=10)

    toggle_button = tk.Button(info_frame, text="切り替え", command=toggle_date_mode,bg=button_bg_color,fg="black")
    toggle_button.grid(row=0, column=9, padx=10)
    #Tao search Box

    search_label = tk.Label(info_frame, text="カーバ記号")
    search_label.grid(row=1, column=0, padx=10)
    search_entry = tk.Entry(info_frame, width=20)
    search_entry.grid(row=1, column=1, padx=10)
    search_button = tk.Button(info_frame, text="検索", command=lambda: search_data(search_entry.get()),bg=button_bg_color,fg="black")
    search_button.grid(row=1, column=2, padx=10)

    def search_data(keyword):
        for item in tree.get_children():
            tree.delete(item)
        # Dữ liệu mẫu để minh họa
        sample_data = [
            ("3D6", "2ndBL", 41),
            ("3D6", "2ndBR", 40),
            ("3D6", "2ndCL", 41),
            ("3D6", "2ndCR", 41),
            ("908", "部位5", 130),
            ("LM", "部位6", 160),
        ]
        filtered_data = [item for item in sample_data if keyword in item[0]]
        for row in filtered_data:
            tree.insert("", tk.END, values=row)


    # Tạo bảng dữ liệu
    columns = ("カバー記号", "部位", "生産数", "先日在庫","出荷数","本日在庫")

    tree_frame = tk.Frame(root)
    tree_frame.pack(expand=True, fill="both")

    scrollbar = ttk.Scrollbar(tree_frame)
    scrollbar.pack(side="right", fill="y")

    tree = ttk.Treeview(
        tree_frame,
        columns=columns,
        show="headings",
        selectmode="browse",
        yscrollcommand=scrollbar.set,
        height=10,
    )
    tree.pack(expand=True, fill="both")
    scrollbar.config(command=tree.yview)

    # Treeview columns
    tree.column("#0", anchor="w", width=120)
    tree.column(columns[0], anchor="w", width=120)
    tree.column(columns[1], anchor="w", width=120)
    tree.column(columns[2], anchor="w", width=120)
    tree.column(columns[3], anchor="w", width=120)
    tree.column(columns[4], anchor="w", width=120)



    # style = ttk.Style()
    # style.configure("Treeview", font=('Calibri', 11), rowheight=25, background="white", foreground="black", spacing1=5, spacing2=5)
    # style.configure("Treeview.Heading", font=('Calibri', 11, 'bold'))
    # style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])
    # style.configure("Treeview", separator=20)

    # tree = ttk.Treeview(root, columns=columns, show="headings", height=10, style="Treeview")


    def update_car_types(event):
        selected_factory = factory_type_combobox.get()
        if selected_factory == "平子":
            car_types = ["CRO", "SPO", "SPT", "STY","全部"]
        elif selected_factory == "明治川":
            car_types = ["908", "LM","全部"]
        else:
            car_types = []  # Trường hợp nhà máy không xác định
        car_type_combobox['values'] = car_types
        if car_types:
            car_type_combobox.current(0)  # Đặt giá trị mặc định cho combobox loại xe

    factory_type_combobox.bind("<<ComboboxSelected>>", update_car_types)
    update_car_types(None)  # Cập nhật loại xe ban đầu


    selected_car_type = car_type_combobox.get()
    selected_factory = factory_type_combobox.get()


    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor=tk.CENTER)

    

    #Ham xu ly API
    def determine_update_params(selected_factory):
        if selected_factory == "平子":
            return 'http://localhost:5006/api/Hirako/get_data', '平子'
        elif selected_factory == "明治川":
            return 'http://localhost:5005/api/Meiji/get_data', '明治川'


    #Ham Load Data

    def load_data():
        selected_car_type = car_type_combobox.get()
       
        selected_date = single_date_entry.get_date()      
        selected_factory = factory_type_combobox.get()
        if(selected_factory=="明治川"):
            file_path = r'C:\Users\0002618\Desktop\API\output_Meijigawa.xlsx'
        else: 
            file_path = r'C:\Users\0002618\Desktop\API\output_Hirako.xlsx'

        df = pd.read_excel(file_path)
        df['車種'] = df['車種'].astype(str)
        df['車種'] = df['車種'].str.strip()
        df['カバー記号'] = df['カバー記号'].str.strip()
        df['部位'] = df['部位'].str.strip()
        df['左右'] = df['左右'].str.strip()   

        df['日付'] = df['日付'].str.strip()            
        df['日付'] = pd.to_datetime(df['日付'], format='%a, %d %b %Y %H:%M:%S %Z')   
        
        if selected_car_type == "908" and selected_factory == "明治川" :
            parts = ["2ndBL", "2ndBR", "2ndCL", "2ndCR"]
            data = []
            count = {}
            cover_symbols = [
                "3D6", "3D5", "4D1", "4D2", "4B1", "4B4", "4C1", "4C4",
                "5D3", "5D4", "5B2", "5B3", "5C2", "5C3", "6D2", "6D3",
                "7D2", "7D3"
            ]
            filtered_df = df[(df['カバー記号'] == '3D6') & (df['部位'] == '2ndB') & (df['左右'] == 'L') & (df['日付'].dt.date == selected_date)]
            count[('3D6', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '3D6') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3D6', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '3D6') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3D6', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '3D6') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3D6', '2ndCR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '3D5') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3D5', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '3D5') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3D5', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '3D5') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3D5', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '3D5') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3D5', '2ndCR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4D1') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4D1', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4D1') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4D1', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4D1') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4D1', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4D1') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4D1', '2ndCR')] = filtered_df.shape[0]       
            filtered_df = df[(df['カバー記号'] == '4D2') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4D2', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4D2') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4D2', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4D2') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4D2', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4D2') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4D2', '2ndCR')] = filtered_df.shape[0] 
            filtered_df = df[(df['カバー記号'] == '5D3') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5D3', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5D3') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5D3', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5D3') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5D3', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5D3') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5D3', '2ndCR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5D4') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5D4', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5D4') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5D4', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5D4') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5D4', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5D4') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5D4', '2ndCR')] = filtered_df.shape[0]                                             
            filtered_df = df[(df['カバー記号'] == '6D2') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6D2', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6D2') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6D2', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6D2') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6D2', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6D2') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6D2', '2ndCR')] = filtered_df.shape[0]                                             
            filtered_df = df[(df['カバー記号'] == '6D3') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6D3', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6D3') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6D3', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6D3') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6D3', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6D3') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6D3', '2ndCR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '7D3') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('7D3', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '7D3') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('7D3', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '7D3') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('7D3', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '7D3') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('7D3', '2ndCR')] = filtered_df.shape[0]  
            filtered_df = df[(df['カバー記号'] == '7D2') & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('7D2', '2ndBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '7D2') & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('7D2', '2ndBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '7D2') & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('7D2', '2ndCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '7D2') & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('7D2', '2ndCR')] = filtered_df.shape[0] 

            filtered_df = df[(df['カバー記号'] == '4C1') & (df['部位'] == '3rdB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4C1', '3rdBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4C1') & (df['部位'] == '3rdB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4C1', '3rdBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4C1') & (df['部位'] == '3rdC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4C1', '3rdCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4C1') & (df['部位'] == '3rdC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4C1', '3rdCR')] = filtered_df.shape[0] 

            filtered_df = df[(df['カバー記号'] == '4C4') & (df['部位'] == '3rdB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4C4', '3rdBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4C4') & (df['部位'] == '3rdB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4C4', '3rdBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4C4') & (df['部位'] == '3rdC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4C4', '3rdCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4C4') & (df['部位'] == '3rdC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4C4', '3rdCR')] = filtered_df.shape[0] 

            filtered_df = df[(df['カバー記号'] == '5C2') & (df['部位'] == '3rdB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5C2', '3rdBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5C2') & (df['部位'] == '3rdB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5C2', '3rdBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5C2') & (df['部位'] == '3rdC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5C2', '3rdCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5C2') & (df['部位'] == '3rdC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5C2', '3rdCR')] = filtered_df.shape[0] 

            filtered_df = df[(df['カバー記号'] == '5C3') & (df['部位'] == '3rdB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5C3', '3rdBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5C3') & (df['部位'] == '3rdB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5C3', '3rdBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5C3') & (df['部位'] == '3rdC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5C3', '3rdCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5C3') & (df['部位'] == '3rdC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5C3', '3rdCR')] = filtered_df.shape[0]    

            filtered_df = df[(df['カバー記号'] == '4B1') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4B1', 'FBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4B1') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4B1', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4B1') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4B1', 'FCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4B1') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4B1', 'FCR')] = filtered_df.shape[0]

            filtered_df = df[(df['カバー記号'] == '4B4') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4B4', 'FBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4B4') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4B4', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4B4') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('4B4', 'FCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '4B4') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('4B4', 'FCR')] = filtered_df.shape[0]    

            filtered_df = df[(df['カバー記号'] == '5B2') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5B2', 'FBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5B2') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5B2', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5B2') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5B2', 'FCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5B2') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5B2', 'FCR')] = filtered_df.shape[0]    

            filtered_df = df[(df['カバー記号'] == '5B3') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5B3', 'FBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5B3') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5B3', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5B3') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('5B3', 'FCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '5B3') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('5B3', 'FCR')] = filtered_df.shape[0]      

            filtered_df = df[(df['カバー記号'] == '6B2') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6B2', 'FBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B2') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6B2', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B2') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6B2', 'FCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B2') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6B2', 'FCR')] = filtered_df.shape[0]                                                               

            filtered_df = df[(df['カバー記号'] == '6B3') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6B3', 'FBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B3') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6B3', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B3') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6B3', 'FCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B3') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6B3', 'FCR')] = filtered_df.shape[0]                                                               


            filtered_df = df[(df['カバー記号'] == '6B4') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6B4', 'FBL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B4') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6B4', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B4') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('6B4', 'FCL')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == '6B4') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('6B4', 'FCR')] = filtered_df.shape[0]



            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                if symbol in ["4B1", "4B4", "5B2", "5B3"]:
                    part = ["FBL", "FBR", "FCL", "FCR"]
                elif symbol in ["4C1", "4C4", "5C2", "5C3"]:
                    part = ["3rdBL", "3rdBR", "3rdCL", "3rdCR"]
                else:
                    part = parts
                for index, p in enumerate(part):
                    if index == 0:
                        data.append((symbol, p, count.get((symbol, p), 0)))
                    else:
                        data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)
            
        elif selected_car_type == "LM" and selected_factory == "明治川": 
            parts = ["2ndBL", "2ndBR", "2ndCL", "2ndCR"]
            data = []
            count = {}
            cover_symbols = ["BR1"]

            filtered_df = df[(df['カバー記号'] == 'BR1') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('BR1', 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['カバー記号'] == 'BR1') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('BR1', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == 'BR1') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('BR1', 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['カバー記号'] == 'BR1') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('BR1', 'FCR')] = filtered_df.shape[0] 
 
            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                if symbol in ["4B1", "4B4", "5B2", "5B3","BR1"]:
                    part = ["FBL", "FBR", "FCL", "FCR"]

                elif symbol in ["4C1", "4C4", "5C2", "5C3"]:
                    part = ["3rdBL", "3rdBR", "3rdCL", "3rdCR"]
                else:
                    part = parts             
                for index, p in enumerate(part):
                    if index == 0:
                        data.append((symbol, p, count.get((symbol, p), 0)))
                    else:
                        data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)

        elif selected_car_type == "CRO" and selected_factory == "平子":
            parts = ["FBL", "FBR", "FCL", "FCR"]
            data = []
            count = {}
            cover_symbols = ["8DG"]

            filtered_df = df[(df['車種'] == '8DG') & (df['カバー記号'] == '8DG') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('8DG', 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['車種'] == '8DG') &(df['カバー記号'] == '8DG') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('8DG', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['車種'] == '8DG')&(df['カバー記号'] == '8DG') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('8DG', 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['車種'] == '8DG')&(df['カバー記号'] == '8DG') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('8DG', 'FCR')] = filtered_df.shape[0] 
 
            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                if symbol in ["8DG"]:
                    part = ["FBL", "FBR", "FCL", "FCR"]
                else:
                    part = parts             
                for index, p in enumerate(part):
                    if index == 0:
                        data.append((symbol, p, count.get((symbol, p), 0)))
                    else:
                        data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)        

        elif selected_car_type == "SPO" and selected_factory == "平子":
            parts = ["RBL", "RBR", "RCL", "RCR"]
            data = []
            count = {}
            cover_symbols = ["3XG"]

            filtered_df = df[(df['車種'] == 'SPO') & (df['カバー記号'] == '3XG') & (df['部位'] == 'RB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'RBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['車種'] == 'SPO') &(df['カバー記号'] == '3XG') & (df['部位'] == 'RB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'RBR')] = filtered_df.shape[0]
            filtered_df = df[(df['車種'] == 'SPO')&(df['カバー記号'] == '3XG') & (df['部位'] == 'RC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'RCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['車種'] == 'SPO')&(df['カバー記号'] == '3XG') & (df['部位'] == 'RC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'RCR')] = filtered_df.shape[0] 
 
            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                if symbol in ["3XG"]:
                    part = ["RBL", "RBR", "RCL", "RCR"]
                else:
                    part = parts             
                for index, p in enumerate(part):
                    if index == 0:
                        data.append((symbol, p, count.get((symbol, p), 0)))
                    else:
                        data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)    
        elif selected_car_type == "SPT" and selected_factory == "平子":
            parts = ["FBL", "FBR", "FCL", "FCR"]
            data = []
            count = {}
            cover_symbols = ["3YM"]

            filtered_df = df[(df['車種'] == 'SPT') & (df['カバー記号'] == '3YM') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3YM', 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['車種'] == 'SPT') &(df['カバー記号'] == '3YM') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3YM', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['車種'] == 'SPT')&(df['カバー記号'] == '3YM') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3YM', 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['車種'] == 'SPT')&(df['カバー記号'] == '3YM') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3YM', 'FCR')] = filtered_df.shape[0] 
 
            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                if symbol in ["3YM"]:
                    part = ["FBL", "FBR", "FCL", "FCR"]
                else:
                    part = parts             
                for index, p in enumerate(part):
                    if index == 0:
                        data.append((symbol, p, count.get((symbol, p), 0)))
                    else:
                        data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)    
        elif selected_car_type == "STY" and selected_factory == "平子":
            parts = ["FBL", "FBR", "FCL", "FCR"]
            data = []
            count = {}
            cover_symbols = ["3XG"]

            filtered_df = df[(df['車種'] == 'STY') & (df['カバー記号'] == '3XG') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['車種'] == 'STY') &(df['カバー記号'] == '3XG') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['車種'] == 'STY')&(df['カバー記号'] == '3XG') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['車種'] == 'STY')&(df['カバー記号'] == '3XG') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'FCR')] = filtered_df.shape[0] 
 
            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                if symbol in ["3XG"]:
                    part = ["FBL", "FBR", "FCL", "FCR"]
                else:
                    part = parts             
                for index, p in enumerate(part):
                    if index == 0:
                        data.append((symbol, p, count.get((symbol, p), 0)))
                    else:
                        data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)    

        elif selected_car_type == "全部" and selected_factory == "平子":

            data = []
            count = {}
            cover_symbols = ["8DG","3XG","3YM"]                      


            filtered_df = df[(df['車種'] == 'STY') & (df['カバー記号'] == '3XG') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['車種'] == 'STY') &(df['カバー記号'] == '3XG') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['車種'] == 'STY')&(df['カバー記号'] == '3XG') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['車種'] == 'STY')&(df['カバー記号'] == '3XG') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'FCR')] = filtered_df.shape[0] 

            filtered_df = df[(df['車種'] == 'SPT') & (df['カバー記号'] == '3YM') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3YM', 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['車種'] == 'SPT') &(df['カバー記号'] == '3YM') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3YM', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['車種'] == 'SPT')&(df['カバー記号'] == '3YM') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3YM', 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['車種'] == 'SPT')&(df['カバー記号'] == '3YM') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3YM', 'FCR')] = filtered_df.shape[0]   


            filtered_df = df[(df['車種'] == 'SPO') & (df['カバー記号'] == '3XG') & (df['部位'] == 'RB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'RBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['車種'] == 'SPO') &(df['カバー記号'] == '3XG') & (df['部位'] == 'RB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'RBR')] = filtered_df.shape[0]
            filtered_df = df[(df['車種'] == 'SPO')&(df['カバー記号'] == '3XG') & (df['部位'] == 'RC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'RCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['車種'] == 'SPO')&(df['カバー記号'] == '3XG') & (df['部位'] == 'RC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3XG', 'RCR')] = filtered_df.shape[0]      



            filtered_df = df[(df['車種'] == '8DG') & (df['カバー記号'] == '8DG') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('8DG', 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['車種'] == '8DG') &(df['カバー記号'] == '8DG') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('8DG', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['車種'] == '8DG')&(df['カバー記号'] == '8DG') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('8DG', 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['車種'] == '8DG')&(df['カバー記号'] == '8DG') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('8DG', 'FCR')] = filtered_df.shape[0]    
 
            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            parent_ids = {} 
            for symbol in cover_symbols:
                if symbol in ["8DG","3YM"]:
                    part = ["FBL", "FBR", "FCL", "FCR"]
                elif symbol in ["3XG"]:
                    part = ["RBL", "RBR", "RCL", "RCR","FBL", "FBR", "FCL", "FCR"]            
                # parent_id = tree.insert("", tk.END, text=symbol, iid=symbol)
                # parent_ids[symbol] = parent_id
                for index, p in enumerate(part):
                    if index == 0:
                        data.append((symbol, p, count.get((symbol, p), 0)))
                    else:
                        data.append(("", p, count.get((symbol, p), 0)))

            for row in data:
                tree.insert("", tk.END, values=row)
            # for row in data:
            #     if row[0] == "":
            #         # Phần tử con, chèn vào phần tử cha tương ứng
            #         parent_symbol = next((symbol for symbol in cover_symbols if (symbol, row[1]) in count), None)
            #         if parent_symbol:
            #             tree.insert(parent_ids[parent_symbol], tk.END, text="", values=row[1:])
            #     else:
            #         # Phần tử cha, đã được chèn từ trước
            #         tree.insert(parent_ids[row[0]], tk.END, text="", values=row[1:])
        else:

            filtered_df = df[df['日付'].dt.date == selected_date]
            for item in tree.get_children():
                tree.delete(item)
            count_label.config(text="合計: 0") 

    #Ham lam moi Data
    def update_data(api_url,factory_name):
        selected_factory = factory_type_combobox.get()
        url = api_url

        try:
            # Gửi yêu cầu API và nhận dữ liệu JSON
            response = requests.get(url)
            data = response.json()

            # Chuyển đổi dữ liệu JSON thành DataFrame của pandas
            df = pd.DataFrame(data)

            # Lưu DataFrame vào một tệp Excel
            input_file = 'Input.xlsx'

            if os.path.exists(input_file):
        # Nếu tệp đã tồn tại, mở tệp và xóa tất cả dữ liệu
                workbook = load_workbook(input_file)
                sheet = workbook.active
                sheet.delete_rows(1, sheet.max_row)
                workbook.save(input_file)
            df.to_excel(input_file, index=False)
            print("Dữ liệu đã được lưu vào tệp Excel")

        except Exception as e:
            print(f"Lỗi khi lấy dữ liệu từ API: {e}")

        # Chuỗi kết nối đến SQL Server
        connection_string = (
            'DRIVER={ODBC Driver 17 for SQL Server};'
            'SERVER=ARC42618\\MEJIGAWA;'  # Thêm một dấu gạch chéo ngược bổ sung
            'DATABASE=KENSA;'
            'Trusted_Connection=yes;'
        )

        try:
            # Tạo kết nối đến SQL Server
            conn = pyodbc.connect(connection_string)
            cursor = conn.cursor()

            cursor.execute("TRUNCATE TABLE KENSA;")
            conn.commit()
            print("Dữ liệu trong bảng KENSA đã được reset thành công")
            # Load workbook
            wb = load_workbook(input_file)
            sheet = wb.active

            # Tạo DataFrame từ dữ liệu Excel
            df = pd.DataFrame(sheet.values, columns=[cell.value for cell in sheet[1]])
            df = df.iloc[1:] 
            df['日付'] = df['Date'] 
            df['工場'] = factory_name
            # Phân tích trường QR-Data
            df['カバー記号'] = df['QR-Data'].str[:6]
            df['カバー記号'] = df['カバー記号'].str.replace('-', '')
            df['部位'] = df['QR-Data'].str[8:12]
            df['部位'] = df['部位'].str.replace('-', '')
            df['左右'] = df['QR-Data'].str[12:13]
            df['連番'] = df['QR-Data'].str[13:16]
            df['車種'] = df['QR-Data'].str[16:19]
            df['車種'] = df['車種'].str.replace('-', '')
            df['縫製日'] = df['QR-Data'].str[19:25]
            df['備考'] = df['QR-Data'].str[25:]

            # Xóa cột QR-Data

            df = df.drop(columns=['QR-Data'])

            # Lưu dữ liệu vào cơ sở dữ liệu
            for index, row in df.iterrows():
                cursor.execute("""
                    INSERT INTO KENSA (カバー記号, 部位, 左右, 車種, 日付, 縫製日, 連番, 備考, 工場)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    row['カバー記号'], row['部位'], row['左右'], row['車種'], row['日付'],
                    row['縫製日'], row['連番'], row['備考'], row['工場']
                )

            # Commit các thay đổi vào cơ sở dữ liệu
            conn.commit()
            print("Dữ liệu đã được lưu vào cơ sở dữ liệu thành công")
            if(selected_factory=="明治川"):
                output_file = 'output_Meijigawa.xlsx'
            else:
                output_file = 'output_Hirako.xlsx'

            # In tất cả bảng ra file Excel
            query = "SELECT * FROM KENSA"
            df = pd.read_sql(query, conn)
            # output_file = 'output_from_sql.xlsx'
            if os.path.exists(output_file):
                try:
                    # Nếu tệp đã tồn tại, mở tệp và xóa tất cả dữ liệu
                    workbook1 = load_workbook(output_file)
                    sheet = workbook1.active
                    sheet.delete_rows(1, sheet.max_row)
                    workbook1.save(output_file)
                except Exception as e:
                    # Nếu có lỗi khi mở tệp, xóa tệp cũ
                    os.remove(output_file)
                    print(f"Tệp bị lỗi đã được xóa: {e}")

            df.to_excel(output_file, index=False)
            print(f"Dữ liệu đã được lưu vào {output_file}")

            # Xóa bảng dữ liệu (nếu cần)
            
            # Đóng kết nối
            conn.close()

        except pyodbc.Error as ex:
            sqlstate = ex.args[1]
            print(f"Lỗi kết nối đến cơ sở dữ liệu: {sqlstate}")        
        messagebox.showinfo("通知", "データが更新されました!")



    #Doan code them 2 nut 
    button_frame = tk.Frame(root)
    button_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)

    # Tạo nút 'データを読込む' và đặt vào trong button_frame
    load_data_button = tk.Button(button_frame, text="データを読込む", command=load_data,bg=button_bg_color,fg="black")
    load_data_button.pack(side=tk.LEFT, padx=5, pady=10)
# button.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

    # Tạo nút 'データを更新' và đặt vào trong button_frame
    update_data_button = tk.Button(button_frame, text="データを更新", command=lambda: update_data(*determine_update_params(factory_type_combobox.get())),bg=button_bg_color,fg="black")
    update_data_button.pack(side=tk.RIGHT, padx=5)


    count_label = tk.Label(root, text="合計: 0", font=("Helvetica", 12))
    count_label.pack(pady=10)
    tree.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    create_app()